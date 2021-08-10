import openpyxl
# source = r'C:\Users\M7MD\Desktop\Work\Data1.xlsx'
source = r'WB.xlsx'
destination = r'WB2.xlsx'


# this function copy data from source file to the destination file.
def copySheet(source, destination):
    wb = openpyxl.load_workbook(source)    # wb : WorkBook
    ws1 = wb.worksheets[0]    # ws : Work sheet
    wb2 = openpyxl.load_workbook(destination)
    ws2 = wb2.worksheets[0]
    for rows in ws1:
        print('type of ws1:',type(ws1))
        print('ws1', ws1)
        for cells in rows:
            print('type of rows:',type(rows))
            print('rows: ',rows)
            print('type of cells:',type(cells))
            print('cells', cells)
            ws2[cells.coordinate].value = cells.value

    print('source work book sheets name:', wb.sheetnames)

    print('destination work book sheets name:', wb2.sheetnames)
    wb2.save(destination)


    
# def copyCellsInRange(source, destination, max_col, max_row):
#
#     wb1 = openpyxl.load_workbook(source)
#     ws1 = wb1.worksheets[0]
#     wb2 = openpyxl.load_workbook(destination)
#     ws2 = wb2.worksheets[0]
#
#     maxRow = ws1.max_row
#     maxCol = ws1.max_column
#
#     for row in range(1, max_row + 1):
#         for col in range(1, max_col + 1):
#             sourceCell = ws1.cell(row=row, column=col)
#             ws2.cell(row=row, column=col).value = sourceCell.value
#     wb2.save(destination)
#     print('Done!')


def copyCellsInRange2(source, destination, sourceSheetnumber, destinationsheetNumber):
    wb1 = openpyxl.load_workbook(source)
    print(wb1.worksheets)
    ws1 = wb1.worksheets[sourceSheetnumber]
    wb2 = openpyxl.load_workbook(destination)
    ws2 = wb2.worksheets[destinationsheetNumber]

    maxRow = ws1.max_row
    maxCol = ws1.max_column

    for row in range(1, maxRow + 1):
        for col in range(1, maxCol + 1):
            sourceCell = ws1.cell(row=row, column=col)
            print('sourceCell: #', sourceCell.row)
            print(sourceCell.value)
            print(ws2.cell(row=row, column=col).value)
            ws2.cell(row=row, column=col).value = sourceCell.value
    # Save the workbook is required to applied the new changes.
    wb2.save(destination)
    print('Done!')


copyCellsInRange2(source, destination, 0, 1)

print('--------------------------------------------------------------------')

# copySheet(source=sourcse, destination= destination)