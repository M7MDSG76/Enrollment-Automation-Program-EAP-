import openpyxl
import panda
sourcse = r'WB.xlsx'
destination = r'WB2.xlsx'


def copySheet(source, destination):
    wb = openpyxl.load_workbook(source)
    ws1 = wb.worksheets[0]
    wb2 = openpyxl.load_workbook(destination)
    ws2 = wb2.worksheets[0]
    for rows in ws1:
        for cells in rows:
            ws2[cells.coordinate].value = cells.value

    print(ws2['C2'].value)
    wb2.save(destination)

def copyCellsInRange(source, destination):
    dataFrame = []
    wb1 = openpyxl.load_workbook(sourcse)
    ws1 = wb1.worksheets[0]
    for row in ws1:
        for col in row:
            dataFrame.append(ws1[col.coordinate].value)
    
    print(dataFrame)

copyCellsInRange(sourcse, destination)
# copySheet(source=sourcse, destination= destination)