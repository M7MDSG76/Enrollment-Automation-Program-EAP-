import openpyxl
import pandas as pd
import numpy as np

source = r'WB.xlsx'
destination = r'WB2.xlsx'
destination2 = r'dest2.xlsx'


# this function copy whole data from source file to the destination file.
def copySheet(source, destination):
    wb = openpyxl.load_workbook(source)  # wb : WorkBook
    ws1 = wb.worksheets[0]  # ws : Work sheet
    wb2 = openpyxl.load_workbook(destination)
    ws2 = wb2.worksheets[0]
    for rows in ws1:
        print('type of ws1:', type(ws1))
        print('ws1', ws1)
        for cells in rows:
            print('type of rows:', type(rows))
            print('rows: ', rows)
            print('type of cells:', type(cells))
            print('cells', cells)
            print('Cells coordinate: ', ws2[cells.coordinate].value)
            print('Typer of Cells coordinate: ', type(ws2[cells.coordinate].value))
            ws2[cells.coordinate].value = cells.value

    print('source work book sheets name:', wb.sheetnames)

    print('destination work book sheets name:', wb2.sheetnames)
    wb2.save(destination)


# this function requires 1-source file path, 2-destination file path,
# 3-index of source sheet, 4-index of destination sheet
def copyCellsInSheet(source, destination, sourceSheetnumber, destinationSheetNumber):
    # Load Workbook 1
    wb1 = openpyxl.load_workbook(source)
    ws1 = wb1.worksheets[sourceSheetnumber]

    # Load Workbook 2
    wb2 = openpyxl.load_workbook(destination)
    ws2 = wb2.worksheets[destinationSheetNumber]

    # Set the Dimensions of the data
    maxRow = ws1.max_row  # The maximum row index containing data (1-based)
    print('max Row:', maxRow)

    maxCol = ws1.max_column  # The maximum Column index containing data (1-based)
    print('max Col: ', maxCol)

    # Copying the Data from Obj1 to Obj2
    for row in range(1, maxRow + 1):
        for col in range(1, maxCol + 1):
            sourceCell = ws1.cell(row=row, column=col)
            print('sourceCell:\n#', sourceCell.row)

            print('SourceCellValue:\n', sourceCell.value)
            ws2.cell(row=row, column=col).value = sourceCell.value
    # Save the workbook is required to applied the new changes.
    wb2.save(destination)
    print('Done!')


def pandasCopySheet(source, sheetIndex):
    wb = openpyxl.load_workbook(source)
    ws = wb.active
    xl = pd.ExcelFile(source)
    dataFrame = xl.parse(sheetIndex)
    print(dataFrame)
    with pd.ExcelWriter(destination2) as writer:
        dataFrame.to_excel(writer, index=False)





def pandasFilterTable(dataFrame):
    d = pd.DataFrame(dataFrame)
    print('dataFrame:\n', d.head(5))
    print('dataFrame fun: \n', d.info())
    x = [0,1,2]
    print(d.lookup(0, d['Hydrogen Sulfide Awareness']))




wb = openpyxl.load_workbook(source)
ws = wb.active
xl = pd.ExcelFile(destination2)
dataFrame = xl.parse(0)

pandasFilterTable(dataFrame)
# pandasCopySheet(source, 0)


# df = pd.DataFrame(np.array(([1, 2, 3], [4, 5, 6])),
#                   index=['mouse', 'rabbit'],
#                   columns=['one', 'two', 'three'])
#
# print(df)
# copyCellsInSheet(source, destination, 0, 0)
#
# print('--------------------------------------------------------------------')
#
# copySheet(source=source, destination=destination2)
